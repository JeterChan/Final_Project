#模擬按鍵行為用套件
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys 
import requests
import json
import pandas as pd
import openpyxl

#爬蟲用
from bs4 import BeautifulSoup 

#情緒分析用
from snownlp import SnowNLP

import time

#資料夾路徑
#C:\Users\user\OneDrive\桌面\專題


#下面這一串資料是在把通知關掉
options = webdriver.ChromeOptions() 
prefs = {'profile.default_content_setting_values':{'notifications': 2}}
options.add_experimental_option('prefs', prefs)
options.add_argument("disable-infobars")
driver = webdriver.Chrome(executable_path=r"C:\Users\User\python-workspace\專題\chromedriver.exe",chrome_options=options)


def grab_yahoo_usersearch():
    options = webdriver.ChromeOptions()  
    prefs = {'profile.default_content_setting_values':{'notifications': 2}}
    options.add_experimental_option('prefs', prefs)
    options.add_argument("disable-infobars")
    driver = webdriver.Chrome(executable_path=r"C:\Users\User\python-workspace\專題\chromedriver.exe",chrome_options=options)


    driver.get('https://tw.news.yahoo.com/')
    time.sleep(5)

    #找到關鍵字的element 並且將關鍵字輸入
    elem = driver.find_element(By.NAME, "p")
    elem.clear()
    ActionChains(driver).double_click(elem).perform()
    elem.send_keys("MLB 巨人")
    elem.send_keys(Keys.RETURN)
    time.sleep(5)

    driver.execute_script("window.scrollTo(0,document.body.scrollHeight)") 
    s_num = 0
    e_num = 1
    i = 0
    count = 0
    print('載入資料開始...')
    while count<30:
        i = i+1
        elements = driver.find_elements(By.CSS_SELECTOR, '#stream-container-scroll-template > li> div > div > div > div > h3')
        s_num = len(elements)

        elements[s_num - 1].location_once_scrolled_into_view  # 捲動加载資料
        time.sleep(5)  # 延遲1秒

        elements = driver.find_elements(By.CSS_SELECTOR, '#stream-container-scroll-template > li> div > div > div > div > h3')
        e_num = len(elements)
        print('捲動頁面到底部第 %d 次, 前次筆數= %d, 現在筆數= %d' % (i, s_num, e_num))
        count = count+1
    print("載入資料結束...")

    soup = BeautifulSoup(driver.page_source, 'lxml')
    print(soup)
    titles = soup.select('#stream-container-scroll-template > li> div > div > div > div > h3')
    driver.close()
    #將所有搜尋資料輸出至記事本
    file = open(r"C:\Users\User\python-workspace\專題\資料\運動\巨人\MLB 巨人(今天).txt","w",encoding='utf-8')
    count = 1
    for i, title in enumerate(titles):
        if (count%4)!=2:
            print(count)
            print(title.text)
            file.write(title.text.strip()+"\n")
        count+=1
        
    file.close()
grab_yahoo_usersearch()

def check_duplicate(): # 過濾掉資料庫內已經有的
    with open(r"C:\Users\User\python-workspace\專題\資料\運動\巨人\MLB 巨人.txt", 'r',encoding='utf-8') as file:
        lines = file.readlines()
    # 建立一個空的集合來儲存已經出現過的標題
    unique_lines = set()

    with open(r"C:\Users\User\python-workspace\專題\資料\運動\巨人\MLB 巨人(今天).txt", 'r',encoding='utf-8') as file:
        new_lines = file.readlines()

    with open(r"C:\Users\User\python-workspace\專題\資料\運動\巨人\MLB 巨人(今天).txt", 'w',encoding='utf-8') as output_file:
        for line in lines:
            if line not in unique_lines:
                unique_lines.add(line)
            #      output_file.write(line)
            #output_file.write("\n")
        for line in new_lines:
            if line not in unique_lines:
                unique_lines.add(line)
                output_file.write(line)
            else:
                print('重複標題：', line)
check_duplicate()

def copy_to_db(): # 新爬出的內容放進資料庫

    with open(r"C:\Users\User\python-workspace\專題\資料\運動\巨人\MLB 巨人(今天).txt", 'r',encoding='utf-8') as file:
        lines = file.readlines()

    with open(r"C:\Users\User\python-workspace\專題\資料\運動\巨人\MLB 巨人.txt", 'w',encoding='utf-8') as output_file:
        for line in lines:
           output_file.write(line)   
copy_to_db()

def add_to_excel(topic, subtopic):
    # Open the Excel file
    workbook = openpyxl.load_workbook('資料庫.xlsx')
    sheet = workbook.active
    file = open(r"C:\Users\User\python-workspace\專題\資料\運動\巨人\MLB 巨人(今天).txt","r",encoding="utf-8")
    # Find the next available row
    next_row = sheet.max_row + 1

    # Write the data to the worksheet
    for line in file :
            sheet.cell(row=next_row, column=1).value = topic
            sheet.cell(row=next_row, column=2).value = subtopic
            sheet.cell(row=next_row, column=3).value = line
            next_row += 1

    # Save the Excel file
    workbook.save('資料庫.xlsx')

add_to_excel('運動', '巨人')


import openpyxl
from ckip_transformers import __version__
from ckip_transformers.nlp import CkipWordSegmenter, CkipPosTagger, CkipNerChunker

# Initialize drivers
print("Initializing drivers ... WS")
ws_driver = CkipWordSegmenter(model="bert-base", device=-1)
print("Initializing drivers ... POS")
pos_driver = CkipPosTagger(model="bert-base", device=-1)
print("Initializing drivers ... NER")
ner_driver = CkipNerChunker(model="bert-base", device=-1)
print("Initializing drivers ... all done")
print()

def clean(sentence_ws, sentence_pos):
  short_with_pos = []
  short_sentence = []
  stop_pos = set(['Nep', 'Nh']) # 這 2 種詞性不保留
  stop_word = set(['中職']) # 停用詞
  for word_ws, word_pos in zip(sentence_ws, sentence_pos):
    # 只留名詞和動詞
    is_N_or_V = word_pos.startswith("V") or word_pos.startswith("N")
    # 去掉名詞裡的某些詞性
    is_not_stop_pos = word_pos not in stop_pos
    # 去掉"中職"這個詞
    is_not_stop_word = word_ws not in stop_word
    # 只剩一個字的詞也不留
    is_not_one_charactor = not (len(word_ws) == 1)
    # 組成串列
    if is_N_or_V and is_not_stop_pos and is_not_stop_word and is_not_one_charactor:
      short_with_pos.append(f"{word_ws}({word_pos})")
      short_sentence.append(f"{word_ws}")
  return (" ".join(short_sentence), " ".join(short_with_pos))

def main():
    txt_file = open(r"C:\Users\User\python-workspace\專題\資料\運動\中職\中職(今天).txt","r",encoding='utf-8')
    text = []

    for title in txt_file:
      text.append(title)

    ws = ws_driver(text)
    pos = pos_driver(ws)
    ner = ner_driver(text)

    file = open(r"C:\Users\User\python-workspace\專題\資料\運動\中職\ws.txt","w",encoding='utf-8')
    pos_file = open(r"C:\Users\User\python-workspace\專題\資料\運動\中職\pos.txt","w",encoding='utf-8')

    print()
    print('=====')

    for sentence, sentence_ws, sentence_pos, sentence_ner in zip(text, ws, pos, ner):
        print("原文：")
        print(sentence)
        (short, res) = clean(sentence_ws, sentence_pos)
        print("斷詞後：")
        print(short)
        file.write(short+"\n")
        print()
        print("斷詞後+詞性標注：")
        print(res)
        pos_file.write(res+"\n")
        print('=====')
    file.close()
    pos_file.close()
if __name__ == "__main__":
    main()

with open(r"C:\Users\User\python-workspace\專題\資料\運動\中職\ws.txt","r",encoding="utf-8") as ws_hasblank, open(r"C:\Users\User\python-workspace\專題\資料\運動\中職\ws_2.txt","w",encoding="utf-8") as ws_noblank:

  for line in ws_hasblank:
    if line.strip():
      ws_noblank.write(line)
ws_hasblank.close()
ws_noblank.close()

"""# KeyBert"""

from keybert import KeyBERT

ws_file = open(r"C:\Users\User\python-workspace\專題\資料\運動\中職\ws_2.txt","r",encoding="utf-8")
kw_file = open(r"C:\Users\User\python-workspace\專題\資料\運動\中職\kw.txt","w",encoding="utf-8")

kw_model = KeyBERT()
for doc in ws_file:
  keywords = kw_model.extract_keywords(doc,keyphrase_ngram_range=(1,1),use_mmr=True, diversity=0.2,top_n=3)
  #for i in range(10):
  for keyword in keywords: #跑三次，因top_n=3
    kw_file.write(keyword[0]+" ")
  kw_file.write("\n")
#print(keywords)

ws_file.close()
kw_file.close()

def add_kw_to_excel():

  # 開啟現有的Excel檔案
  workbook = openpyxl.load_workbook(r"C:\Users\User\python-workspace\專題\資料庫.xlsx")
  # 選取要寫入的工作表，假設為第一個工作表
  sheet = workbook.active

  kw_file = open(r"C:\Users\User\python-workspace\專題\資料\運動\中職\kw.txt","r",encoding="utf-8")
  # Find the next available row
  next_row = 117

  for line in kw_file:
        sheet.cell(row=next_row, column=5, value=line)
        next_row += 1 


  kw_file.close()
  # 儲存Excel檔案
  workbook.save(r"C:\Users\User\python-workspace\專題\資料庫.xlsx")

add_kw_to_excel()


def word2vec():
  from gensim.models import Word2Vec
  import logging

  # 設定logging
  logging.basicConfig(format='%(asctime)s : %(levelname)s : %(message)s', level=logging.INFO)

  # 載入文本數據
  corpus = []
  with open(r'C:\Users\User\python-workspace\專題\資料\運動\中職\kw.txt', 'r', encoding='utf-8') as f:
      for line in f:
          corpus.append(line.strip())
  #print(len(corpus),'筆資料')
  #print(corpus)

  # 轉換文本數據為tokens列表
  tokens = [sentence.split() for sentence in corpus]

  # 訓練Word2Vec模型
  model = Word2Vec(tokens,vector_size=81,window=5,min_count=10,sg=0,hs=0,negative=10,workers=3,epochs=10,sample=0.05)

  # 儲存模型
  model.save('your_model.model')

  # 加載模型
  loaded_model = Word2Vec.load('your_model.model')

  # 查找相似詞
  similar_words = loaded_model.wv.most_similar("郭天信", topn=5)
  print(similar_words)

#word2vec()
