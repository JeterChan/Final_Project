#模擬按鍵行為用套件
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys 
import requests
import json
import pandas as pd
import openpyxl
from pymongo import MongoClient

#爬蟲用
from bs4 import BeautifulSoup 



#下面這一串資料是在把通知關掉
options = webdriver.ChromeOptions() 
prefs = {'profile.default_content_setting_values':{'notifications': 2}}
options.add_experimental_option('prefs', prefs)
options.add_argument("disable-infobars")
driver = webdriver.Chrome(executable_path=r"C:\Users\User\python-workspace\專題\chromedriver.exe",chrome_options=options)


def grab_yahoo_usersearch(topic,date):
    options = webdriver.ChromeOptions()  
    prefs = {'profile.default_content_setting_values':{'notifications': 2}}
    options.add_experimental_option('prefs', prefs)
    options.add_argument("disable-infobars")
    driver = webdriver.Chrome(executable_path=r"C:\Users\User\python-workspace\專題\chromedriver.exe",chrome_options=options)


    driver.get('https://tw.news.yahoo.com/')
    time.sleep(5)

    #找到關鍵字的element 並且將關鍵字輸入
    elem = driver.find_element(By.NAME, "p")
    elem.clear()
    ActionChains(driver).double_click(elem).perform()
    elem.send_keys(topic)
    elem.send_keys(Keys.RETURN)
    time.sleep(5)

    driver.execute_script("window.scrollTo(0,document.body.scrollHeight)") 
    s_num = 0
    e_num = 1
    i = 0
    count = 0
    print('載入資料開始...')
    while count<30:
        i = i+1
        elements = driver.find_elements(By.CSS_SELECTOR, '#stream-container-scroll-template > li> div > div > div > div > h3')
        s_num = len(elements)

        elements[s_num - 1].location_once_scrolled_into_view  # 捲動加载資料
        time.sleep(5)  # 延遲1秒

        elements = driver.find_elements(By.CSS_SELECTOR, '#stream-container-scroll-template > li> div > div > div > div > h3')
        e_num = len(elements)
        print('捲動頁面到底部第 %d 次, 前次筆數= %d, 現在筆數= %d' % (i, s_num, e_num))
        count = count+1
    print("載入資料結束...")

    soup = BeautifulSoup(driver.page_source, 'lxml')
    print(soup)
    titles = soup.select('#stream-container-scroll-template > li> div > div > div > div > h3')
    driver.close()
    #將所有搜尋資料輸出至記事本
    file = open("C:\\Users\\User\\python-workspace\\專題\\資料\\運動\\" + topic + "\\"+ date + topic + ".txt", "w", encoding='utf-8')
    count = 1
    for i, title in enumerate(titles):
        if (count%4)!=2:
            print(count)
            print(title.text)
            file.write(title.text.strip()+"\n")
        count+=1
        
    file.close()

def check_duplicate(topic,date): # 過濾掉資料庫內已經有的
    with open("C:\\Users\\User\\python-workspace\\專題\\資料\\運動\\" + topic + "\\" + topic + ".txt", 'r',encoding='utf-8') as file:
        lines = file.readlines()
    # 建立一個空的集合來儲存已經出現過的標題
    unique_lines = set()

    with open("C:\\Users\\User\\python-workspace\\專題\\資料\\運動\\" + topic + "\\"+ date  + topic + ".txt", 'r',encoding='utf-8') as file:
        new_lines = file.readlines()

    with open("C:\\Users\\User\\python-workspace\\專題\\資料\\運動\\" + topic + "\\"+ date  + topic + ".txt", 'w',encoding='utf-8') as output_file:
        for line in lines:
            if line not in unique_lines:
                unique_lines.add(line)
        for line in new_lines:
            if line not in unique_lines:
                unique_lines.add(line)
                output_file.write(line)
            else:
                print('重複標題：', line)


def copy_to_db(title):
    # 新爬出的內容放進資料庫
    uri = "mongodb+srv://user1:user1@cluster0.ronm576.mongodb.net/?retryWrites=true&w=majority"

# Create a new client and connect to the server
    client = MongoClient(uri, server_api=ServerApi('1'))

# Send a ping to confirm a successful connection
    try:
        
        client.admin.command('ping')
        print("Pinged your deployment. You successfully connected to MongoDB!")
    # 获取要插入的数据
        data = {
    "title": title
}

# 选择要插入数据的数据库和集合
        db = client["News"]
        collection = db["Sport"]

# 插入数据
        collection.insert_one(data)


    except Exception as e:
        print(e)


def add_to_excel(topic, subtopic,date):
    # Open the Excel file
    workbook = openpyxl.load_workbook('資料庫.xlsx')
    sheet = workbook.active
    file = open("C:\\Users\\User\\python-workspace\\專題\\資料\\運動\\" + subtopic + "\\"+ date  + subtopic + ".txt","r",encoding="utf-8")
    # Find the next available row
    next_row = sheet.max_row + 1

    # Write the data to the worksheet
    for line in file :
            sheet.cell(row=next_row, column=1).value = topic
            sheet.cell(row=next_row, column=2).value = subtopic
            sheet.cell(row=next_row, column=3).value = line
            next_row += 1

    # Save the Excel file
    workbook.save('資料庫.xlsx')

if __name__ == '__main__':
    topics = ['棒球','中職','美職','日職','韓職','中信兄弟','味全龍','統一獅','樂天桃猿','富邦悍將','台鋼雄鷹','洋基','紅襪','光芒','金鶯','藍鳥','守護者','白襪','皇家','老虎','雙城','太空人','運動家','水手','天使','遊騎兵','大都會','勇士','費城人','馬林魚','國民','釀酒人','紅雀','紅人','小熊','海盜','響尾蛇','道奇','落磯','巨人','教士']
    date="5.27"
    for topic in topics:
        grab_yahoo_usersearch(topic,date)
        check_duplicate(topic,date)
        copy_to_db(topic,date)
        add_to_excel('運動',topic,date) 

